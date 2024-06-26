from django.shortcuts import render,redirect
from redminelib import Redmine
from redminelib.exceptions import AuthError
import ssl
from openpyxl import load_workbook
from datetime import datetime, timedelta,date
from . import models
import mysql.connector
from django.http import JsonResponse
import requests
from .decorators import check_login_session, initialize_redmine
from django.contrib import messages
from django.http import HttpResponse
import os
from django.core.exceptions import ObjectDoesNotExist

ssl._create_default_https_context = ssl._create_unverified_context
# key = "a435e1173a8238a1fb5fd6d07bc8042c901abbfd"

# ACCESS TO REDMINE, USED IN DECORATORS
# redmine = Redmine('https://redmine.greenfieldsdairy.com/redmine', key=key, requests={'verify': False,})
# redmine = Redmine('https://redmine.greenfieldsdairy.com/redmine', username=username, password=password, requests={'verify': False,})

# STORE USERS AS EMPTY 
users = []


# LOGIN
def login(request):
    if request.method == "GET":
        # FIND SESSION FROM RECENT LOGIN
        if request.session.get('username') and request.session.get('password'):
            return redirect('index')
        else:
            return render(request, 'login.html')
    else:
        username = request.POST['username']
        password = request.POST['password']
        try:
            redmine = Redmine('https://redmine.greenfieldsdairy.com/redmine', 
                                    username=username, 
                                    password=password, 
                                    requests={'verify': False})
            # TEST AUTH ACCESS
            redmine.user.get('current')

            # STORE INTO SESSION FOR CACHING
            request.session['username'] = username
            request.session['password'] = password
            
            users = redmine.user.get('current')
            user = users.id

            # COMES FROM DEF REGISTER, USED TO DELETE AND RETRIEVE DB TO GET AN UPDATED DATA
            try:
                connection = mysql.connector.connect(
                    host='10.58.1.2',
                    port='3307',
                    database='bitnami_redmine',
                    user='report',
                    password='mokondo12'
                )
                if connection.is_connected():
                    print('Connected to MySQL database')
                cursor = connection.cursor()
                query = "SELECT users.id, users.firstname, users.lastname, address FROM users JOIN email_addresses ON users.id = email_addresses.user_id"
                cursor.execute(query)
                rows = cursor.fetchall()
                models.user.objects.all().delete()
                for row in rows:
                    id, firstname, lastname, email = row
                    try:
                        models.user.objects.create(id=id,name=firstname+ ' ' +lastname, email=email)
                    except:
                        print('not exists')
            except mysql.connector.Error as e:
                print(f'Error connecting to MySQL database: {e}')

            finally:
                # Close cursor and connection
                if 'cursor' in locals():
                    cursor.close()
                if 'connection' in locals() and connection.is_connected():
                    connection.close()
                    print('Connection to MySQL database closed')

            # FIND A SETTINGS THAT USER HAS, IF NOT THEN CREATE USING A DEFAULT TEMPLATE
            try:
                models.settings.objects.get(user=user)
            except ObjectDoesNotExist:
                models.settings.objects.create(user=user)
            return redirect('index')
        
        # WRONG USERNAME/PASSWORD HANDLER
        except AuthError:
            error_message = "Invalid username or password. Please try again!"
            messages.error(request, error_message)
            return render(request, 'login.html', {'error_message': error_message})

# LOGOUT
def logout(request):
    # DELETE LOGIN SESSION TO LOGOUT
    del request.session['username']
    del request.session['password']
    return redirect ('login')

# DASHBOARD
@check_login_session
@initialize_redmine
def index(request, redmine):
    users = redmine.user.get('current')
    user = users.id
    listproject = []
    selected_tasks = []

    # SELECT ALL PROJECT THAT AUTHORIZED BY ME AND APPEND IT ON LISTPROJECT
    try:
        connection = mysql.connector.connect(
            host='10.58.1.2',
            port='3307',
            database='bitnami_redmine',
            user='report',
            password='mokondo12'
        )
        if connection.is_connected():
            print('Connected to MySQL database')

        cursor = connection.cursor()
        query1 = "SELECT projects.id, projects.name FROM projects JOIN members ON projects.id = members.project_id WHERE members.user_id = %s AND projects.status !=5 ORDER BY projects.name ASC "
        today_date = date.today() + timedelta(days=3)
        start_date = date(today_date.year, today_date.month, today_date.day)
        author_id = user
        cursor.execute(query1, (author_id,))
        rows1 = cursor.fetchall()
        if rows1:
            for row in rows1:
                list = {}
                list['id'] = row[0]
                list['name'] = row[1]
                listproject.append(list)
        len_project = len(listproject)

        # SELECT REMAINING TASK THAT ASSIGNED BY ME LAST 5 WITH THE CLOSEST DUE DATE
        query2 = "SELECT issues.id, issues.subject, issues.due_date, issues.project_id, projects.name FROM issues JOIN projects ON issues.project_id = projects.id WHERE assigned_to_id = %s AND projects.status !=5 AND due_date <= %s"
    
        cursor.execute(query2, (author_id, start_date))
        rows2 = cursor.fetchall()
        len_task = len(rows2)
        if rows2:
            for row in rows2:
                task_info = {
                'id' : row[0],
                'name' : row[1],
                'project' : row[4],
                'due_date' : row[2]
                }
                selected_tasks.append(task_info)
                
                # ONLY 3 TASKS
                if len(selected_tasks) == 3:
                    break
        else:
            print("No rows found matching the criteria.")

    except mysql.connector.Error as e:
        print(f'Error connecting to MySQL database: {e}')

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals() and connection.is_connected():
            connection.close()
            print('Connection to MySQL database closed')

    # CLICK IN REMAINING TASK
    if request.method == "POST":
        go = request.POST['go']
        return redirect ('listissue', id=str(go))

    return render(request,'index.html',{
        'selected_tasks' : selected_tasks,
        'len_task' : len_task,
        'name' : users['firstname'] +' ' + users['lastname'],
        'listproject' : listproject,
        'len_project' : len_project
    }) 

# PROJECT
@check_login_session
@initialize_redmine
def listproject(request, redmine):
    listproject = []
    users = redmine.user.get('current')
    user = users.id

     # SELECT ALL PROJECTS WHERE USER IS A MEMBER AND APPEND INTO LIST PROJECT
    try:
        connection = mysql.connector.connect(
            host='10.58.1.2',
            port='3307',
            database='bitnami_redmine',
            user='report',
            password='mokondo12'
        )
        if connection.is_connected():
            print('Connected to MySQL database')

        cursor = connection.cursor()
        query = "SELECT projects.name, projects.identifier FROM projects JOIN members ON projects.id = members.project_id WHERE members.user_id = %s AND projects.status !=5 ORDER BY projects.name ASC"
        cursor.execute(query, (user,))
        rows = cursor.fetchall()
        if rows:
            for row in rows:
                # 0 IS PROJECT_NAME, 1 IS PROJECT_ID
                list = []
                list.append(row[0])
                list.append(row[1])
                listproject.append(list)
        else:
            print("No rows found matching the criteria.")

    except mysql.connector.Error as e:
        print(f'Error connecting to MySQL database: {e}')

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals() and connection.is_connected():
            connection.close()
            print('Connection to MySQL database closed')
    return render(request, 'listproject.html',{
        'listproject' : listproject
    })

@check_login_session
@initialize_redmine
def updateproject(request,id, redmine):
    update = redmine.project.get(id)
    if request.method == "GET":
        return render(request, 'updateproject.html',{
            'update' : update,
            'listproject' : listproject
        })
    else:
        update.description = request.POST['description']
        update.is_public = eval(request.POST['is_public'])
        update.save()
        return redirect('listproject')
    

@check_login_session
@initialize_redmine
def newproject(request, redmine):
    listproject = []
    new = redmine.project.new()
    users = redmine.user.get('current')
    user = users.id
    # SELECT ALL PROJECTS THAT USER IS A MEMBER FOR ALL PROJECTS DROPDOWN
    try:
        connection = mysql.connector.connect(
            host='10.58.1.2',
            port='3307',
            database='bitnami_redmine',
            user='report',
            password='mokondo12'
        )
        if connection.is_connected():
            print('Connected to MySQL database')

        cursor = connection.cursor()
        query = "SELECT projects.id, projects.name FROM projects JOIN members ON projects.id = members.project_id WHERE members.user_id = %s AND projects.status !=5 ORDER BY projects.name ASC"
        cursor.execute(query, (user,))
        rows = cursor.fetchall()
        if rows:
            for row in rows:
                list = {}
                list['id'] = row[0]
                list['name'] = row[1]
                listproject.append(list)
        else:
            print("No rows found matching the criteria.")

    except mysql.connector.Error as e:
        print(f'Error connecting to MySQL database: {e}')

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals() and connection.is_connected():
            connection.close()
            print('Connection to MySQL database closed')

    if request.method == "GET":
        return render (request, 'newproject.html',{
            'listproject' : listproject
        })
    elif request.method == "POST":
        new.name = request.POST['name']

        # HANDLER REMOVE SPACE IN NAME FOR PROJECT IDENTIFIER
        if " " in request.POST['name']:
            name = request.POST['name'].replace(" ", "")
            name = name.lower()
            new.identifier = name
        else:
            print(request.POST['name'])
            new.identifier = request.POST['name'].lower()
        
        new.description = request.POST['description']
        new.is_public = eval(request.POST['is_public'])

        # HANDLER IF USER INPUT A PARENT PROJECT
        if request.POST['parent'] == "kosong":
            pass
        else:
            new.parent_id = request.POST['parent']

        new.save()
        return redirect ('confirmation', name=request.POST['name'])
    
@check_login_session
@initialize_redmine
def confirmation (request, name, redmine):
    
    if request.method == 'POST':
        users = redmine.user.get('current')
        user = users.id
        get = models.settings.objects.get(user=user)
        # STATUS, DEPT, USER, AND PRIORITY IS STORED IN LOCAL DB TO FASTEN THE DROPDOWN LOAD
        allstatus = models.status.objects.all()
        alldept = models.dept.objects.all()
        alluser = models.user.objects.all()
        allpriority = models.priority.objects.all()

        forprint = list()
        if 'upload' in request.POST:
            # LOAD XLSX
            files = request.FILES['filea']
            wb = load_workbook(files)
            ws = wb.active
            
            # START GANTT CHART ROW
            start_row = get.start_row

            # STORED COLUMN VALUES

            def get_date_range_for_week(year, week_start, week_end):
                if week_start is None or week_start < 1:
                    return []
                first_day = datetime(year, 1, 1)
                def calculate_week_dates(week_number):
                    offset = (week_number - 1) * 7
                    start_date = first_day + timedelta(days=offset)
                    end_date = start_date + timedelta(days=4)
                    return start_date, end_date
                if week_end is None:
                    start_date, end_date = calculate_week_dates(week_start)
                    return [(start_date, end_date)] 
                elif week_start is not None and week_end is not None and week_end >= week_start:
                    start_date = calculate_week_dates(week_start)[0]
                    end_date = calculate_week_dates(week_end)[1]
                    return [(start_date, end_date)]
                else:
                    return []
                
            def find_col_with_filled_color(ws, row_index):
                filled_cells = []
                # 9 STATE THE COLUMN GANTT CHART STARTED
                for col_index in range(get.gannt_start_column, ws.max_column + 1):
                    cell = ws.cell(row=row_index, column=col_index)
                    if isinstance(cell.fill.fgColor.theme, int) or (cell.fill.fgColor.rgb != 'FFFF0000' and  cell.fill.fgColor.rgb !='00000000'):
                        findweek = ws.cell(get.week_number_row, col_index).value
                        filled_cells.append(findweek)

                return filled_cells if filled_cells else None
            
            all = {}
            all['col_start'] = []
            for i in range(1, get.end_column - get.start_column):
                all[f'col{i}_values'] = []
            all['col_end'] = []

            # ITERATE COLUMN 2 AS A BASE

            for row in ws.iter_rows(min_row=start_row,min_col=get.start_column,max_col=get.start_column, values_only=True):
                cell_value = row[0] if row and row[0] is not None else None
                all['col_start'].append(cell_value)  
            # ITERATE ROW FOR EVERY COLUMN TO GET THE VALUES
            for i in range(1, get.end_column - get.start_column):
                for row in ws.iter_rows(min_row=start_row, min_col=get.start_column+i, max_col=get.start_column+i, values_only=True):
                    cell_value = row[0] if row and row[0] is not None else None
                    all[f'col{i}_values'].append(cell_value)
            for row in ws.iter_rows(min_row=start_row, min_col=get.end_column, max_col=get.end_column, values_only=True):
                cell_value = row[0] if row and row[0] is not None else None
                all['col_end'].append(cell_value)        
            
            
            tasks_data = []
            current_task = None
           
        # Dictionary to keep track of the current parent task at each column level
            current_tasks = {key: None for key in all.keys()}

            for idx, value in enumerate(all['col_start']):
                if value is not None:
                    # Main Task
                    print('masuk')
                    p = find_col_with_filled_color(ws, idx + start_row)
                    if p is not None:
                        if len(p) > 1:
                            week_start = min(p)
                            week_end = max(p)
                        else:
                            week_start = min(p)
                            week_end = None

                        get_ranges = get_date_range_for_week(2024, week_start, week_end)
                        for start_date, end_date in get_ranges:
                            start_date = start_date.strftime('%Y-%m-%d')
                            end_date = end_date.strftime('%Y-%m-%d')
                    else:
                        week_start = None
                        week_end = None
                        start_date = week_start
                        end_date = week_end

                    findpic = ws.cell(idx + start_row, get.pic_column).value
                    if findpic is not None:
                        findpic = findpic.upper()
                    findperson = ws.cell(idx + start_row, get.email_column).value
                    if findperson is not None:
                        if ',' in findperson:
                            findperson = findperson.split(',')
                        findperson = findperson

                    current_task = {
                        'name': value,
                        'start_date': start_date,
                        'due_date': end_date,
                        'pic': findpic,
                        'person': findperson,
                        'subtasks': []
                    }
                    tasks_data.append(current_task)

                    # Set the current task for col_start level
                    current_tasks['col_start'] = current_task

                    # Reset the current task for all subsequent columns
                    for key in list(current_tasks.keys())[1:]:
                        current_tasks[key] = None

                else:
                    # Iterate through remaining columns to find subtasks
                    for key in list(all.keys())[1:]:
                        col_values = all[key]
                        subtask_value = col_values[idx]
                        if subtask_value is not None:
                            print(key)
                            print(subtask_value)
                            p = find_col_with_filled_color(ws, idx + start_row)
                            if p is not None:
                                if len(p) > 1:
                                    week_start = min(p)
                                    week_end = max(p)
                                else:
                                    week_start = min(p)
                                    week_end = None

                                get_ranges = get_date_range_for_week(2024, week_start, week_end)
                                for start_date, end_date in get_ranges:
                                    start_date = start_date.strftime('%Y-%m-%d')
                                    end_date = end_date.strftime('%Y-%m-%d')
                            else:
                                week_start = None
                                week_end = None
                                start_date = week_start
                                end_date = week_end

                            findpic = ws.cell(idx + start_row, get.pic_column).value
                            if findpic is not None:
                                findpic = findpic.upper()
                            findperson = ws.cell(idx + start_row, get.email_column).value
                            if findperson is not None:
                                if ',' in findperson:
                                    findperson = findperson.split(',')
                                findperson = findperson

                            subtask = {
                                'name': subtask_value,
                                'pic': findpic,
                                'start_date': start_date,
                                'due_date': end_date,
                                'person': findperson,
                                'subtasks': []
                            }

                            # Find the appropriate parent task for the current subtask
                            for parent_key in reversed(list(current_tasks.keys())[:list(current_tasks.keys()).index(key)]):
                                if current_tasks[parent_key] is not None:
                                    if 'subtasks' not in current_tasks[parent_key]:
                                        current_tasks[parent_key]['subtasks'] = []
                                    current_tasks[parent_key]['subtasks'].append(subtask)
                                    break

                            # Set the current task for the current column
                            current_tasks[key] = subtask

                            # Reset the current task for all subsequent columns
                            for subsequent_key in list(current_tasks.keys())[list(current_tasks.keys()).index(key) + 1:]:
                                current_tasks[subsequent_key] = None

            

            print(tasks_data)
            # IT IS USED TO ATTACH A PARENT NAME KEYS FOR A SUBTASKS BASED ON A SUBTASKS THAT CONTAINED IN PARENT
            def process_tasks(tasks, parent_name=None):
                for task in tasks:
                    task['Parent'] = parent_name
                    forprint.append(task)
                    if 'subtasks' in task and task['subtasks']:
                        process_tasks(task['subtasks'], parent_name=task['name'])
                    if 'subtasks' in task:
                        del task['subtasks']
            
            process_tasks(tasks_data)
            print(forprint)
            user = redmine.user.get('current')

            # print(forprint)
            # SESSION THE DATA TO BE PASSED INTO AFTER POST LOGIC

            request.session['forprint'] = forprint
            request.session['name'] = name

            return render (request, 'confirmation.html', {
                'name' : name,
                'tasks_data' : forprint,
                'alluser' : alluser,
                'user' : user,
                'allstatus' : allstatus,
                'alldept' : alldept,
                'allpriority' : allpriority
            })
        else:
            # IN THIS SECTION, THE DATA IS BEING CREATED INTO REDMINE DB
            # GET THE SESSION
            forprint = request.session.get('forprint',[])
            name = request.session.get('name',[])
            def create_issue(item):
                # THE ['name'] IS USE TO HANDLE A MULTIPLE FORM FOR A MULTIPLE TASK THAT WILL BE PROCCESSED
                
                id = name
                form_name = f"name_{item['name']}"
                form_description = f"description_{item['name']}"
                form_start_date = f"start_date_{item['name']}"
                form_due_date = f"due_date_{item['name']}"
                form_assigned_to = f"assigned_to_{item['name']}"
                form_department = f"department_{item['name']}[]"
                form_responsible = f"responsible_{item['name']}[]"
                form_status= f"status_{item['name']}"
                form_priority = f"priority_{item['name']}"
                form_estimated_hours = f"estimated_hours_{item['name']}"
                form_done_ratio = f"done_ratio_{item['name']}"
                form_parent = f"parenttask_{item['name']}"
                parent = request.POST.get(form_parent)
                subject = request.POST.get(form_name)
                start_date = request.POST.get(form_start_date)
                due_date = request.POST.get(form_due_date)
                assigned_to = request.POST.get(form_assigned_to)
                description = request.POST.get(form_description)
                status = request.POST.get(form_status)
                priority = request.POST.get(form_priority)
                estimated_hours = request.POST.get(form_estimated_hours)
                done_ratio = request.POST.get(form_done_ratio)
                department = request.POST.getlist(form_department)
                responsible = request.POST.getlist(form_responsible)
                nama_parent = form_parent.split("_")
                nama_parent = nama_parent[1]
                create = redmine.issue.new()

                # IF THERE IS A PARENT ON TASK
                def get_latest(redmine, project_id):
                    latest = redmine.issue.filter(project_id = project_id, sort='created_on:desc', limit=1)
                    if latest:
                        return latest[0].id
                    return None
                def get_latest_issue_id(redmine, project_id):
                    # Retrieve the latest issue created in the project
                    latest_issue = redmine.issue.filter(project_id=project_id, sort='created_on:desc', limit=1)
                    if latest_issue:
                        return latest_issue[0].id
                    return None

                if parent is not None:
                    # Get the latest issue ID in the project
                    latest_issue_id = get_latest_issue_id(redmine, id)
                    
                    if latest_issue_id is not None:
                        parent_issue_ids = redmine.issue.filter(project_id=id, subject=parent)
                        
                        if parent_issue_ids:
                            # Find the issue with the ID closest to the latest issue ID
                            closest_issue = min(parent_issue_ids, key=lambda issue: abs(issue.id - latest_issue_id))
                            
                            # Set the parent issue ID to the closest ID found
                            create.parent_issue_id = closest_issue.id

                    
                    # parent_issue_id = redmine.issue.filter(project_id = id, subject=parent)
                    # print('nama', subject)
                    # for i in parent_issue_id:
                    #     print('opsi', i, i.id)
                    # if parent_issue_id:
                    #     largest = max(issue.id for issue in parent_issue_id)
                    #     create.parent_issue_id = largest
                    #     print('pilih',largest)
                else:
                    pass
                create.project_id = id
                create.subject=subject
                create.description = description
                create.assigned_to_id = assigned_to
                create.start_date=start_date
                create.due_date=due_date
                create.status=status
                create.priority=priority
                create.estimated_hours=estimated_hours
                create.done_ratio=done_ratio
                create.custom_fields=[{'id':11, 'value':department}]
                create.custom_fields=[{'id':4, 'value':responsible}]
                create.save()

            # STORED INTO DIFFERENT CONDITION ITEMS, SO THE PARENT IS THE ONE WHO REGISTERED FIRST
            first_condition_items = []
            second_condition_items = []
            last_condition_items = []
            all_responsible_users = []


            # LOGIC TO ASSIGN A DIFFERENT CONDITION
            responsible2 = []
            for item in forprint:
                responsible_users = request.POST.getlist(f'responsible_{item["name"]}[]')
                responsible2.append(responsible_users)
                if item.get('Parent') is None:
                    first_condition_items.append(item)
                # elif item.get('subtasks') is not None and item.get('Parent') is not None:
                else:
                    second_condition_items.append(item)
                # elif item.get('subtasks') is None and item.get('Parent') is not None:
                #     last_condition_items.append(item)
                # else:
                #     print('stillexists')
            print('first',first_condition_items)
            print('second',second_condition_items)
            print('last',last_condition_items)
            flat = [item for sublist in responsible2 for item in sublist]
            all_responsible_users = list(set(flat))

            # ADD A RESPONSIBLE AS A MEMBER FIRST
            if all_responsible_users:
                for user_id in all_responsible_users:
                    try:
                        new_membership = redmine.project_membership.new()
                        new_membership.project_id = name
                        new_membership.user_id = user_id
                        new_membership.role_ids = [6]  
                        new_membership.save()
                    except:
                        pass
            
            # CREATE_ISSUE BASED ON THE CONDITION
            for item in first_condition_items:
                create_issue(item)
            for item in second_condition_items:
                create_issue(item)
            for item in last_condition_items:
                create_issue(item)
            
            return redirect('listproject')
        
    return render(request, 'confirmation.html', {\
        })

# ISSUE
@check_login_session
@initialize_redmine
def listissue(request,id, redmine):
    global users
    listissue = []
    try:
        connection = mysql.connector.connect(
            host='10.58.1.2',
            port='3307',
            database='bitnami_redmine',
            user='report',
            password='mokondo12'
        )
        if connection.is_connected():
            print('Connected to MySQL database')

        # SELECT ALL ISSUES THAT ASSIGNED BY ME ON CERTAIN PROJECT
            
        cursor = connection.cursor()
        query = "SELECT issues.id, issues.subject, issues.status_id, users.firstname, users.lastname, issues.start_date, issues.due_date FROM issues JOIN projects ON issues.project_id = projects.id JOIN users ON issues.assigned_to_id = users.id WHERE projects.identifier = %s AND projects.status !=5"
       
        cursor.execute(query, (id,))
        rows = cursor.fetchall()
        if rows:
            for row in rows:
                list = []
                list.append(row[0])
                list.append(row[1])
                p = models.status.objects.get(id=row[2])
                list.append(p.name)
                try:
                    list.append(row[3] + " " + row[4])
                except:
                    pass
                list.append(row[5])
                list.append(row[6])
                # 1 IS ISSUE_NAME, 2, IS ISSUE_STATUS, 3 + 4 IS FIRST_NAME + LAST_NAME, 5 IS START_DATE, 6 IS DUE_DATE
                listissue.append(list)
        
        else:
            print("No rows found matching the criteria.")

    except mysql.connector.Error as e:
        print(f'Error connecting to MySQL database: {e}')

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals() and connection.is_connected():
            connection.close()
            print('Connection to MySQL database closed')
    
    
    
    if request.method == "GET":
        return render(request, 'listissue.html',{
        'listissue' : listissue,
        'id' : id
    })

@check_login_session
@initialize_redmine
def export(request, id, redmine):
    tes = redmine.issue.filter(project_id=id)
    file_path = tes.export('csv', savepath='../redmine/', columns='all')
    with open(file_path, 'rb') as file:
        csv_data = file.read()

    # DELETE AFTER BEING READ
    os.remove(file_path)

    # DOWNLOAD POPUP
    response = HttpResponse(csv_data, content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=redmine_data_{id}.csv'
    return response

@check_login_session
@initialize_redmine
def listdetails(request,id, redmine):
    get = redmine.issue.get(id)
    
    dict = {}
    listresponsible = []
    listdepartment = []
    dict['subject'] = get.subject
    dict['id'] = get.id
    dict['description'] = get.description
    dict['status'] = get.status
    dict['priority'] = get.priority
    dict['assigned_to'] = '-'
    
    # GET RESPONSIBLE LIST AND JOIN WITH , IF THERE IS MULTIPLE
    try:
        if get.custom_fields[0]['value']:
            for i in get.custom_fields[0]['value']:
                user = models.user.objects.get(id=int(i))
                listresponsible.append(user.name)
            listresponsible = ', '.join(listresponsible)
            dict['responsible'] = listresponsible
        else:
            dict['responsible'] = '-'
    except:
        pass

    try:
        if dict['assigned_to']:
            dict['assigned_to'] = get.assigned_to
    except:
        pass

    try:
        if get.custom_fields[7]['value']:
            for i in get.custom_fields[7]['value']:
                    listdepartment.append(i)
            listdepartment = ', '.join(listdepartment)
            dict['department'] = listdepartment
        else:
            dict['department'] = '-'
    except:
        pass

    # FIND RELATION IF AVAILABLE
    relation  = redmine.issue_relation.filter(issue_id = id)
    if relation:
        list = {}
        issue_id = []
        relation_type = []
        issue_to_id = []
        for i in relation:
            issue_id.append(i['issue_id'])
            issue_to_id.append(i['issue_to_id'])
            relation_type.append(i['relation_type'].capitalize())
        list['issue_id'] = issue_id
        list['relation_type'] = relation_type
        list['issue_to_id'] = issue_to_id
        dict['relation'] = list
        relation_str = ""
        for issue_id, issue_to_id, relation_type in zip(issue_id, issue_to_id, relation_type):
            relation_str += f"<a href='/listdetails/{issue_id}'>#{issue_id}</a> {relation_type} to <a href='/listdetails/{issue_to_id}'>#{issue_to_id}</a>, "
        relation_str = relation_str[:-2]
        dict['relation'] = relation_str
    else:
        dict['relation'] = '-'

    
    dict['start_date'] = get.start_date
    dict['due_date'] = get.due_date
    dict['estimated_hours'] = get.estimated_hours
    dict['done_ratio'] = get.done_ratio    

   
    return render(request, 'listdetails.html',{
        'dict' : dict
    })

@check_login_session
@initialize_redmine
def newissue(request, redmine):
    user = redmine.user.get('current')
    alldept = models.dept.objects.all()
    allstatus = models.status.objects.all()
    allpriority = models.priority.objects.all()
    alluser = models.user.objects.all()
    listproject = []

    # SELECT AVAILABLE PROJECT THAT USER IS A MEMBER TO BEING PARENT DROPDOWN
    try:
        connection = mysql.connector.connect(
            host='10.58.1.2',
            port='3307',
            database='bitnami_redmine',
            user='report',
            password='mokondo12'
        )
        if connection.is_connected():
            print('Connected to MySQL database')

        
        cursor = connection.cursor()
        query = "SELECT projects.id, projects.name FROM projects JOIN members ON projects.id = members.project_id WHERE members.user_id = %s AND projects.status !=5 ORDER BY projects.name ASC "
        author_id = user.id
        cursor.execute(query, (author_id,))
        rows = cursor.fetchall()
        if rows:
            for row in rows:
                list = {}
                list['id'] = row[0]
                list['name'] = row[1]
                listproject.append(list)
        
        else:
            print("No rows found matching the criteria.")

    except mysql.connector.Error as e:
        print(f'Error connecting to MySQL database: {e}')

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals() and connection.is_connected():
            connection.close()
            print('Connection to MySQL database closed')


    if request.method == "GET":
        return render(request, 'newissue.html',{
            'all' : all,
            'user' : user,
            'listproject' : listproject,
            'allstatus' : allstatus,
            'alldept' : alldept,
            'allpriority' : allpriority,
            'alluser' : alluser
        })
    else:
        new = redmine.issue.new()
        id = request.POST['id']

        status = request.POST['status']
        getstatus = models.status.objects.get(id=status)
        
        priority = request.POST['priority']
        getpriority = models.priority.objects.get(id=priority)
        
        
        responsible = request.POST.getlist('responsible[]')
        department = request.POST.getlist('department[]')
        

        if responsible:
                for user_id in responsible:
                    try:
                        new_membership = redmine.project_membership.new()
                        new_membership.project_id = id
                        new_membership.user_id = user_id
                        new_membership.role_ids = [6]
                        new_membership.save()
                    except:
                        pass

        new.project_id = id
        new.subject = request.POST['subject']
        new.description = request.POST['description']
        new.status_id = getstatus.id
        new.priority_id = getpriority.id
        new.assigned_to_id = request.POST['assigned_to']
        new.custom_fields=[{'id':4, 'value':responsible}]
        new.custom_fields=[{'id':11, 'value':department}]
        new.estimated_hours = request.POST['estimated_hours']
        new.start_date = request.POST['start_date']
        new.due_date = request.POST['due_date']
        new.done_ratio = request.POST['done_ratio']
        new.save()
        return redirect('listproject')
    
def handler404(request, exception, template_name='404.html'):
    response = render(template_name)
    response.status_code = 404
    return response

def handler500(request,template_name='500.html'):
    response = render(template_name)
    response.status_code = 500
    return response

@check_login_session
@initialize_redmine
def updateissue(request,id, redmine):
    listdept = []
    listresponsible = []
    update = redmine.issue.get(id)
    start_date = datetime.strftime(update.start_date, '%Y-%m-%d')
    due_date = datetime.strftime(update.due_date, '%Y-%m-%d')
    alluser = models.user.objects.all()
    allstatus = models.status.objects.all()
    allpriority = models.priority.objects.all()
    alldept = models.dept.objects.all()
    dept = update.custom_fields
    for i in update.custom_fields[0]['value']:
        listresponsible.append(int(i))
    for i in dept[7]['value']:
        listdept.append(i)
    
    if request.method == "GET":
        return render(request, 'updateissue.html',{
            'id' : id,
            'update' : update,
            'alluser' : alluser,
            'listdept' : listdept,
            'allstatus' : allstatus,
            'allpriority' : allpriority,
            'alldept' : alldept,
            'start_date' : start_date,
            'due_date' : due_date,
            'listresponsible' : listresponsible
        })
    else:
        
        status = request.POST['status']
        getstatus = models.status.objects.get(id=status)
        priority = request.POST['priority']
        getpriority = models.priority.objects.get(id=priority)

        responsible = request.POST.getlist('responsible[]')
        department = request.POST.getlist('department[]')

        get = redmine.issue.get(id)
        project = get.project
        if responsible:
                for user_id in responsible:
                    try:
                        new_membership = redmine.project_membership.new()
                        new_membership.project_id = project
                        new_membership.user_id = user_id
                        new_membership.role_ids = [6]  
                        new_membership.save()
                    except:
                        pass

        update.subject = request.POST['subject']
        update.status_id = getstatus.id
        update.description = request.POST['description']
        
        update.custom_fields=[{'id':11,'value':department}]
        update.start_date = request.POST['start_date']
        update.due_date = request.POST['due_date']
        update.priority_id = getpriority.id
        update.assigned_to_id = request.POST['assigned_to']
        update.custom_fields=[{'id':4, 'value':responsible}]
        update.estimated_hours = request.POST['estimated_hours']
        update.done_ratio = request.POST['done_ratio']
        update.save()
        return redirect('listdetails',id=id)
@check_login_session
@initialize_redmine    
def deleteissue(request,id, redmine):
    delete = redmine.issue.get(id)
    p = delete['project']
    delete.delete()
    return redirect('listissue', id=p)

@check_login_session
@initialize_redmine
def addrelations(request,id, redmine):
    if request.method == "GET":
        get = redmine.issue.filter(project_id = id)
        return render(request, 'addrelation.html',{
            'get':get
        })
    else:
        target = redmine.issue_relation.new()
        target.issue_id = request.POST['issue_id']
        target.issue_to_id = request.POST['issue_to_id']
        target.relation_type = request.POST['relation_type']
        if request.POST['relation_type'] == "precedes" or "follows":
            target.delay = request.POST['delay']
        target.save()
        return redirect ('listissue', id=id)

# CREATED AS A MANUAL DATA UPDATE, ACCESS IT THROUGH /REGISTER
@check_login_session
def register(request):
    try:
        connection = mysql.connector.connect(
            host='10.58.1.2',
            port='3307',
            database='bitnami_redmine',
            user='report',
            password='mokondo12'
        )
        if connection.is_connected():
            print('Connected to MySQL database')
        cursor = connection.cursor()
        query = "SELECT users.id, users.firstname, users.lastname, address FROM users JOIN email_addresses ON users.id = email_addresses.user_id"
        cursor.execute(query)
        rows = cursor.fetchall()
        models.user.objects.all().delete()
        for row in rows:
            id, firstname, lastname, email = row
            try:
                models.user.objects.create(id=id,name=firstname+ ' ' +lastname, email=email)
            except:
                print('not exists')
    except mysql.connector.Error as e:
        print(f'Error connecting to MySQL database: {e}')

    finally:
        # Close cursor and connection
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals() and connection.is_connected():
            connection.close()
            print('Connection to MySQL database closed')

    return redirect('index')



# THIS FUNCTION IS USED TO GET AN AVAILABLE EMAIL TO BEING AUTO SCHEDULE IT EVERY MORNING
# ITS FOR MANUAL TEST, ACCESSS IT THROUGH /send_email
@check_login_session
def send_email(request):
    if request.method == "GET":
        try:
            connection = mysql.connector.connect(
                host='10.58.1.2',
                port='3307',
                database='bitnami_redmine',
                user='report',
                password='mokondo12'
            )
            if connection.is_connected():
                print('Connected to MySQL database')

            # Create a cursor object
            cursor = connection.cursor()

            # Define the SQL query
            query = [
                "SELECT issues.id, issues.subject, issues.project_id, projects.name, issues.start_date, issues.due_date, issues.author_id FROM issues JOIN projects ON issues.project_id = projects.id WHERE start_date = %s AND projects.status !=5 AND issues.status_id !=5",
                "SELECT issues.id, issues.subject, issues.project_id, projects.name, issues.start_date, issues.due_date, issues.author_id FROM issues JOIN projects ON issues.project_id = projects.id WHERE start_date <= %s AND %s <= due_date AND projects.status !=5 AND issues.status_id !=5",
                "SELECT issues.id, issues.subject, issues.project_id, projects.name, issues.start_date, issues.due_date, issues.author_id FROM issues JOIN projects ON issues.project_id = projects.id WHERE due_date = %s AND projects.status !=5 AND issues.status_id !=5",

            ]
            today_date = date.today()
            start_date = date(today_date.year, today_date.month, today_date.day)
            params_list = [
                (start_date,),
                (start_date, start_date,),
                (start_date,)
            ]
            results = []
            for queries, params in zip(query, params_list):
                cursor.execute(queries,params)
                rows=cursor.fetchall()
                results.append(rows)
            print(results)
            startselected_tasks = results[0]
            between_tasks = results[1]
            dueselected_tasks = results[2]
            
            startselected_tasks = [item for item in startselected_tasks]
            between_tasks = [item for item in between_tasks]
            dueselected_tasks = [item for item in dueselected_tasks]
            def convert(i):
               return [
                    {
                        "id": row[0],
                        "subject": row[1],
                        "project_id": row[2],
                        "project_name": row[3],
                        "start_date": row[4],
                        "due_date": row[5],
                        "author_id" : row[6]
                        
                    }
                    for row in i
                ]
            startselected_tasks = convert(startselected_tasks)
            between_tasks = convert(between_tasks)
            dueselected_tasks = convert(dueselected_tasks)

            print('start today :', startselected_tasks)
            print('on progress :',between_tasks)
            print('due today :',dueselected_tasks)
            
            # query = "SELECT value FROM custom_values WHERE custom_field_id=4 AND customized_id = idissue"
            # query = "SELECT column_name FROM information_schema.columns WHERE table_schema='bitnami_redmine' AND table_name = 'custom_values'"
            # query = "SELECT table_schema, table_name FROM information_schema.tables WHERE table_type = 'BASE TABLE'"

         
        except mysql.connector.Error as e:
            print(f'Error connecting to MySQL database: {e}')

        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'connection' in locals() and connection.is_connected():
                connection.close()
                print('Connection to MySQL database closed')

        def email(list):
            clean = []
            try:
                connection = mysql.connector.connect(
                    host='10.58.1.2',
                    port='3307',
                    database='bitnami_redmine',
                    user='report',
                    password='mokondo12'
                )
                if connection.is_connected():
                    print('Connected to MySQL database')

                for item in list:
                    # Create a cursor object
                    cursor = connection.cursor()
                    responsible_id = []
                    query = "SELECT value FROM custom_values WHERE custom_field_id=4 AND customized_id = %s"
                    params = (item['id'],)
                    cursor.execute(query, params)
                    rows = cursor.fetchall()
                    if rows:
                        for row in rows:
                            if row == ('',):
                                continue
                            else:
                                print('awal')
                                responsible_id.append(row[0])
                        print(responsible_id)
                    
                    if responsible_id:
                        for i in responsible_id:
                            query3 = "SELECT a.firstname, a.lastname, b.address FROM users AS a JOIN email_addresses AS b ON a.id = b.user_id WHERE a.id=%s"
                            params3 = (i,)
                            cursor.execute(query3,params3)
                            rows = cursor.fetchall()
                            if rows:
                                for row in rows:
                                    responsible = {
                                        'responsible_name' : row[0] + " " + row[1],
                                        'responsible_email' : row[2]
                                    }
                                print(responsible)

                            query2 = "SELECT a.firstname, a.lastname, b.address FROM users AS a JOIN email_addresses as b ON a.id = b.user_id WHERE a.id=%s"
                            params2 = (item['author_id'],)
                            cursor.execute(query2, params2)
                            rows = cursor.fetchall()
                            if rows:
                                for row in rows:
                                    print(row)
                                    author = {
                                        'author_name' : row[0] + " " + row[1],
                                        'author_email' : row[2]
                                    }
                            
                                user_info = {
                                    'id': item['id'],
                                    'subject': item['subject'],
                                    'project_id': item['project_id'],
                                    'project_name': item['project_name'],
                                    'start_date': item['start_date'],
                                    'due_date': item['due_date'],
                                    'author_email' : author['author_email'],
                                    'author_name' : author['author_name'],
                                    'name': responsible['responsible_name'],
                                    'email': responsible['responsible_email']
                                    }
                                clean.append(user_info)
                                
                                
                                if list is startselected_tasks:
                                    print("starts")
                                    body = f"""
                                    Dear {responsible['responsible_name']},

                                    You have a task that started today and  will be due on {item['due_date']} from {item['project_name']} Project
                                    with task subject : {item['subject']}.

                                    View more on : http://redmine.greenfieldsdairy.com/redmine/issues/{item['id']}

                                    Regards,
                                    {author['author_name']}
                                    """
                        
                                elif list is between_tasks:
                                    print("between")
                                    body = f"""
                                    Dear {responsible['responsible_name']},

                                    You still have a running task that started from {item['start_date']} and will be due on {item['due_date']}
                                    from {item['project_name']} Project with task subject : {item['subject']}.

                                    View more on : http://redmine.greenfieldsdairy.com/redmine/issues/{item['id']}

                                    Regards,
                                    {author['author_name']}
                                    """
                                
                                else:
                                    print("last")
                                    body = f"""
                                    
                                    Dear {responsible['responsible_name']},

                                    You have a task that will be due today on {item['due_date']} from {item['project_name']} Project
                                    with task subject : {item['subject']}

                                    View more on : http://redmine.greenfieldsdairy.com/redmine/issues/{item['id']}

                                    Regards,
                                    {author['author_name']}

                                    """
                          

                        for i in clean:
                            email_api = "http://10.24.7.70:3333/send-email"
                            payload = {
                                "to": [i['email']],
                                "cc" : [i['author_email']],
                                "subject": f"#{i['id']} [{i['subject']}] Task Reminder",
                                "body": body
                            }

                            print(payload)
                            print('akhir')
                            # response = requests.post(email_api, json = payload)

                            # if response.status_code == 200:
                            #     print("Email sent successfully.")
                            # else:
                            #     print(f"Failed to send email. Status code: {response.status_code}")
                            #     print(response.text)  # Print the response content for debugging
                     
                    else:
                        print('gaada')
                        break
                    responsible_id.clear()
                    author.clear()
                    responsible.clear()
            except mysql.connector.Error as e:
                print(f'Error connecting to MySQL database: {e}')

            finally:
                if 'cursor' in locals():
                    cursor.close()
                if 'connection' in locals() and connection.is_connected():
                    connection.close()
                    print('Connection to MySQL database closed')

        email(startselected_tasks)
        email(between_tasks)
        email(dueselected_tasks)
        
        return redirect('index')

@check_login_session
@initialize_redmine
def settings(request,redmine):
    users = redmine.user.get('current')
    user = users.id
    if request.method == "GET":
        try:
            get = models.settings.objects.get(user=user)
            
            dict = {
                'start_row' : get.start_row,
                'gannt_start_column' : get.gannt_start_column,
                'week_number_row' : get.week_number_row,
                'start_column' : get.start_column,
                'end_column' : get.end_column,
                'pic_column' : get.pic_column,
                'email_column' : get.email_column
            }
        except ObjectDoesNotExist:
            create = models.settings.objects.create(user=user)
            get = create
            dict={
                'start_row' : get.start_row,
                'gannt_start_column' : get.gannt_start_column,
                'week_number_row' : get.week_number_row,
                'start_column' : get.start_column,
                'end_column' : get.end_column,
                'pic_column' : get.pic_column,
                'email_column' : get.email_column
            }
        return render(request, 'settings.html',{
    'dict' : dict

    })
    else:

        get = models.settings.objects.get(user=user)
        get.gannt_start_column = request.POST['gannt_start_column']
        get.week_number_row = request.POST['week_number_row']
        get.start_row = request.POST['start_row']
        get.start_column = request.POST['start_column']
        get.end_column = request.POST['end_column']
        get.pic_column = request.POST['pic_column']
        get.email_column = request.POST['email_column']
        get.save()
        return redirect('settings')
    
    
